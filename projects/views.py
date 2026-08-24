from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.http import HttpResponseRedirect
from django.utils.timezone import now
from django.db.models import Q

from core.permissions import Level2RequiredMixin, Level3RequiredMixin, Level4RequiredMixin
from .models import (
    Project, ProjectCategory, ProjectAllocation, ProjectLifecycleStage, 
    ProjectFee, FeeType, ProjectMonitoringLog, ProjectMonitoringImage,
    SubcontractorPaymentTranche, ProjectActivityLog
)
from .forms import (
    ProjectForm, ProjectAllocationForm, ProjectLifecycleStageForm, 
    ProjectFeeFormSet, ProjectCategoryForm, FeeTypeForm, ProjectMonitoringLogForm,
    ProjectMonitoringLogGlobalForm, SubcontractorPaymentTrancheForm
)
from .utils import log_project_activity

class ProjectListView(LoginRequiredMixin, ListView):
    model = Project
    template_name = 'projects/project_list.html'
    context_object_name = 'projects'
    paginate_by = 30

    def get_queryset(self):
        qs = super().get_queryset().select_related('category')
        q = self.request.GET.get('q', '').strip()
        year = self.request.GET.get('year', '').strip()
        mda = self.request.GET.get('mda', '').strip()
        category = self.request.GET.get('category', '').strip()
        project_type = self.request.GET.get('project_type', '').strip()
        awarded = self.request.GET.get('awarded', '').strip().lower()

        if q:
            qs = qs.filter(
                Q(project_code__icontains=q) | Q(project_name__icontains=q)
            )
        if year:
            qs = qs.filter(created_at__year=year)
        if mda:
            qs = qs.filter(mda__icontains=mda)
        if category:
            if category.isdigit():
                qs = qs.filter(category_id=category)
            else:
                if category.upper() == 'CONSTRUCTION':
                    qs = qs.filter(Q(category__name__iexact='CONSTRUCTION') | Q(category__name__icontains='Civil'))
                else:
                    qs = qs.filter(category__name__iexact=category)
        if project_type:
            qs = qs.filter(project_type=project_type)
        if awarded in ['awarded', 'yes', '1', 'true']:
            qs = qs.filter(
                Q(actual_contract_amount__gt=0) |
                Q(current_phase__in=['POST_AWARD', 'EXECUTION']) |
                (Q(award_letter_and_boq__isnull=False) & ~Q(award_letter_and_boq=''))
            )
        elif awarded in ['pre_award', 'no', '0', 'false']:
            qs = qs.filter(
                Q(actual_contract_amount=0) &
                Q(current_phase='PRE_AWARD') &
                (Q(award_letter_and_boq__isnull=True) | Q(award_letter_and_boq=''))
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass current filter values back so the form stays populated
        selected_cat = self.request.GET.get('category', '').strip()
        cat_display = ''
        if selected_cat:
            if selected_cat.isdigit():
                c = ProjectCategory.objects.filter(pk=selected_cat).first()
                cat_display = c.name if c else selected_cat
            else:
                cat_display = selected_cat

        context['q'] = self.request.GET.get('q', '')
        context['selected_year'] = self.request.GET.get('year', '')
        context['selected_mda'] = self.request.GET.get('mda', '')
        context['selected_category'] = selected_cat
        context['selected_category_display'] = cat_display
        context['selected_type'] = self.request.GET.get('project_type', '')
        context['selected_awarded'] = self.request.GET.get('awarded', '')

        # Preserve filter parameters for pagination links
        get_copy = self.request.GET.copy()
        if 'page' in get_copy:
            del get_copy['page']
        context['querystring'] = get_copy.urlencode()

        if context.get('is_paginated'):
            page_obj = context['page_obj']
            context['page_range'] = list(page_obj.paginator.get_elided_page_range(
                number=page_obj.number, on_each_side=1, on_ends=1
            ))
        else:
            context['page_range'] = [1]

        # Build distinct filter option lists from the full table
        context['year_choices'] = (
            Project.objects.dates('created_at', 'year', order='DESC')
        )
        context['mda_choices'] = (
            Project.objects.values_list('mda', flat=True).distinct().order_by('mda')
        )
        context['category_choices'] = ProjectCategory.objects.all().order_by('name')
        context['type_choices'] = Project.PROJECT_TYPE_CHOICES
        return context


class ProjectCreateView(Level3RequiredMixin, CreateView):
    model = Project
    form_class = ProjectForm
    template_name = 'projects/project_form.html'
    success_url = reverse_lazy('projects:project_list')

    def get_initial(self):
        initial = super().get_initial()
        parent_id = self.request.GET.get('parent_project')
        if parent_id:
            try:
                parent = Project.objects.get(pk=parent_id)
                initial.update({
                    'parent_project': parent.pk,
                    'project_code': parent.project_code,
                    'project_name': parent.project_name,
                    'mda': parent.mda,
                    'location': parent.location,
                    'category': parent.category_id,
                    'project_type': parent.project_type,
                    'execution_mode': parent.execution_mode,
                    'lot': parent.lot,
                    'budget_amount': parent.budget_amount,
                    'actual_contract_amount': parent.actual_contract_amount,
                })
                # Calculate remaining percentage
                existing_parts_sum = sum(p.part_percentage for p in parent.project_parts.all())
                initial['part_percentage'] = max(0, 100 - existing_parts_sum)
            except Project.DoesNotExist:
                pass
        return initial

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['fee_formset'] = ProjectFeeFormSet(self.request.POST)
        else:
            data['fee_formset'] = ProjectFeeFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        fee_formset = context['fee_formset']
        if fee_formset.is_valid():
            self.object = form.save()
            fee_formset.instance = self.object
            fee_formset.save()
            
            log_project_activity(
                project=self.object,
                user=self.request.user,
                action_type='CREATE',
                title=f"Project Created ({self.object.project_code})",
                description=f"Created new project '{self.object.project_name}' for MDA {self.object.mda}."
            )
            messages.success(self.request, "Project created successfully!")
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class ProjectUpdateView(Level3RequiredMixin, UpdateView):
    model = Project
    form_class = ProjectForm
    template_name = 'projects/project_form.html'

    def get_success_url(self):
        return reverse_lazy('projects:project_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['fee_formset'] = ProjectFeeFormSet(self.request.POST, instance=self.object)
        else:
            data['fee_formset'] = ProjectFeeFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        fee_formset = context['fee_formset']
        if fee_formset.is_valid():
            # Track field diffs before saving
            old_proj = Project.objects.get(pk=self.object.pk)
            changes = {}
            tracked_fields = [
                ('project_name', 'Title'),
                ('mda', 'MDA'),
                ('lot', 'Lot'),
                ('budget_amount', 'Budget Amount'),
                ('actual_contract_amount', 'Contract Amount'),
                ('execution_level_percentage', 'Execution %'),
                ('technical_status', 'Technical Status'),
                ('payment_status', 'Payment Status'),
                ('current_phase', 'Current Phase'),
            ]
            for field_name, label in tracked_fields:
                old_val = getattr(old_proj, field_name)
                new_val = form.cleaned_data.get(field_name)
                if old_val != new_val:
                    changes[label] = {
                        'from': str(old_val if old_val is not None else '-'),
                        'to': str(new_val if new_val is not None else '-')
                    }

            self.object = form.save()
            fee_formset.instance = self.object
            fee_formset.save()

            diff_desc_parts = [f"{lbl}: {d['from']} → {d['to']}" for lbl, d in changes.items()]
            desc = ", ".join(diff_desc_parts) if diff_desc_parts else "Updated project specifications."

            log_project_activity(
                project=self.object,
                user=self.request.user,
                action_type='UPDATE',
                title=f"Updated Project Specs ({self.object.project_code})",
                description=desc,
                changes_json=changes
            )

            messages.success(self.request, "Project updated successfully!")
            return redirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))

class ProjectDeleteView(Level3RequiredMixin, DeleteView):
    model = Project
    template_name = 'projects/confirm_delete.html'
    success_url = reverse_lazy('projects:project_list')

class ProjectDetailView(LoginRequiredMixin, DetailView):
    model = Project
    template_name = 'projects/project_detail.html'
    context_object_name = 'project'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Fetch subcontractor allocations with prefetched tranches
        context['allocations'] = ProjectAllocation.objects.filter(project=self.object).select_related('subcontractor').prefetch_related('payment_tranches')
        context['tranche_form'] = SubcontractorPaymentTrancheForm()
        # Fetch lifecycle stages ordered by sequence_order
        lifecycle_stages = self.object.lifecycle_stages.all()
        context['lifecycle_stages'] = lifecycle_stages
        
        # Group lifecycle stages by phases
        phase1 = []
        phase2 = []
        phase3 = []
        
        # Thresholds based on project type (from signals.py logic)
        if self.object.project_type == 'CONSTRUCTION':
            phase2_limit = 25
        else:
            phase2_limit = 19
            
        for stage in lifecycle_stages:
            if stage.sequence_order <= 17:
                phase1.append(stage)
            elif stage.sequence_order <= phase2_limit:
                phase2.append(stage)
            else:
                phase3.append(stage)
                
        context['phase1_stages'] = phase1
        context['phase2_stages'] = phase2
        context['phase3_stages'] = phase3

        # Form for adding allocation in modal/page
        context['allocation_form'] = ProjectAllocationForm()
        # Calculate totals
        total_incurred_cost = sum(stage.incurred_cost for stage in lifecycle_stages)
        context['total_incurred_cost'] = total_incurred_cost
        
        # Fetch completion percentage based on completed stages vs total stages
        total_stages = len(lifecycle_stages)
        completed_stages = sum(1 for s in lifecycle_stages if s.is_completed)
        if total_stages > 0:
            context['calculated_completion_percentage'] = int((completed_stages / total_stages) * 100)
        else:
            context['calculated_completion_percentage'] = 0
            
        # Fetch site monitoring logs and their images
        context['monitoring_logs'] = self.object.monitoring_logs.all().select_related('reported_by').prefetch_related('images')
        context['monitoring_form'] = ProjectMonitoringLogForm()
        context['recent_activities'] = self.object.activity_logs.select_related('user').all()[:25]
        return context

class ProjectAllocationCreateView(Level3RequiredMixin, View):
    def post(self, request, project_pk):
        project = get_object_or_404(Project, pk=project_pk)
        form = ProjectAllocationForm(request.POST, request.FILES)
        if form.is_valid():
            allocation = form.save(commit=False)
            allocation.project = project
            try:
                allocation.save()
                log_project_activity(
                    project=project,
                    user=request.user,
                    action_type='SUBCONTRACTOR',
                    title=f"Subcontractor Allocated: {allocation.subcontractor.name}",
                    description=f"Agreed Amount: ₦{allocation.amount_agreed_with_supplier_contractor:,.2f} | Advance: ₦{allocation.advance_received_by_supplier_contractor:,.2f}"
                )
                messages.success(request, "Subcontractor allocated successfully!")
            except Exception as e:
                messages.error(request, f"Failed to allocate subcontractor: {e}")
        else:
            messages.error(request, "Invalid form submission.")
        return redirect('projects:project_detail', pk=project.pk)

class ProjectAllocationUpdateView(Level3RequiredMixin, UpdateView):
    model = ProjectAllocation
    form_class = ProjectAllocationForm
    template_name = 'projects/allocation_form.html'

    def get_success_url(self):
        return reverse_lazy('projects:project_detail', kwargs={'pk': self.object.project.pk})

    def form_valid(self, form):
        messages.success(self.request, "Allocation updated successfully!")
        return super().form_valid(form)

class ProjectAllocationDeleteView(Level3RequiredMixin, DeleteView):
    model = ProjectAllocation
    template_name = 'confirm_delete.html'

    def get_success_url(self):
        return reverse_lazy('projects:project_detail', kwargs={'pk': self.object.project.pk})

class UpdateLifecycleStageView(Level3RequiredMixin, View):
    def post(self, request, pk):
        stage = get_object_or_404(ProjectLifecycleStage, pk=pk)
        is_completed = request.POST.get('is_completed') == 'true'
        notes = request.POST.get('notes_or_updates', '')
        incurred_cost = request.POST.get('incurred_cost', '0.00')

        stage.is_completed = is_completed
        stage.notes_or_updates = notes
        try:
            stage.incurred_cost = float(incurred_cost)
        except ValueError:
            pass

        if is_completed and not stage.completed_date:
            stage.completed_date = now().date()
        elif not is_completed:
            stage.completed_date = None

        stage.save()
        log_project_activity(
            project=stage.project,
            user=request.user,
            action_type='STAGE',
            title=f"Stage Progress: {stage.stage_name}",
            description=f"Status: {'Completed' if is_completed else 'In Progress'} | Cost: ₦{stage.incurred_cost:,.2f}" + (f" | Notes: {notes}" if notes else "")
        )
        messages.success(request, f"Updated step: {stage.stage_name}")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))


class ProjectSettingsView(Level3RequiredMixin, View):

    def get_context_data(self, request, cat_form=None, fee_form=None, active_tab='categories'):
        edit_cat_id = request.GET.get('edit_cat')
        edit_fee_id = request.GET.get('edit_fee')
        
        if cat_form is None:
            cat_instance = get_object_or_404(ProjectCategory, pk=edit_cat_id) if edit_cat_id else None
            cat_form = ProjectCategoryForm(instance=cat_instance)
            
        if fee_form is None:
            fee_instance = get_object_or_404(FeeType, pk=edit_fee_id) if edit_fee_id else None
            fee_form = FeeTypeForm(instance=fee_instance)
            
        return {
            'categories': ProjectCategory.objects.all(),
            'fee_types': FeeType.objects.all(),
            'cat_form': cat_form,
            'fee_form': fee_form,
            'is_editing_cat': edit_cat_id is not None,
            'is_editing_fee': edit_fee_id is not None,
            'edit_cat_id': edit_cat_id,
            'edit_fee_id': edit_fee_id,
            'active_tab': active_tab,
        }

    def get(self, request):
        edit_fee_id = request.GET.get('edit_fee')
        active_tab = 'feetypes' if edit_fee_id else 'categories'
        return render(request, 'projects/settings.html', self.get_context_data(request, active_tab=active_tab))

    def post(self, request):
        action = request.POST.get('action')
        if action == 'save_category':
            edit_cat_id = request.GET.get('edit_cat')
            cat_instance = get_object_or_404(ProjectCategory, pk=edit_cat_id) if edit_cat_id else None
            form = ProjectCategoryForm(request.POST, instance=cat_instance)
            if form.is_valid():
                form.save()
                messages.success(request, "Project category saved successfully.")
                return redirect('projects:settings')
            else:
                messages.error(request, "Failed to save category. Please correct the errors.")
                context = self.get_context_data(request, cat_form=form, active_tab='categories')
                return render(request, 'projects/settings.html', context)
                
        elif action == 'save_feetype':
            edit_fee_id = request.GET.get('edit_fee')
            fee_instance = get_object_or_404(FeeType, pk=edit_fee_id) if edit_fee_id else None
            form = FeeTypeForm(request.POST, instance=fee_instance)
            if form.is_valid():
                form.save()
                messages.success(request, "Fee type saved successfully.")
                return redirect('projects:settings')
            else:
                messages.error(request, "Failed to save fee type. Please correct the errors.")
                context = self.get_context_data(request, fee_form=form, active_tab='feetypes')
                return render(request, 'projects/settings.html', context)

        elif action == 'delete_category':
            cat_id = request.POST.get('cat_id')
            category = get_object_or_404(ProjectCategory, pk=cat_id)
            if category.project_set.exists():
                messages.error(request, f"Cannot delete category '{category.name}' because it is assigned to projects.")
            else:
                category.delete()
                messages.success(request, "Category deleted successfully.")
            return redirect('projects:settings')

        elif action == 'delete_feetype':
            fee_id = request.POST.get('fee_id')
            feetype = get_object_or_404(FeeType, pk=fee_id)
            if feetype.projectfee_set.exists():
                messages.error(request, f"Cannot delete fee type '{feetype.name}' because it is assigned to projects.")
            else:
                feetype.delete()
                messages.success(request, "Fee type deleted successfully.")
            return redirect('projects:settings')


class ProjectMonitoringLogCreateView(Level2RequiredMixin, View):
    def post(self, request, project_pk):
        project = get_object_or_404(Project, pk=project_pk)
        form = ProjectMonitoringLogForm(request.POST)
        if form.is_valid():
            log_entry = form.save(commit=False)
            log_entry.project = project
            log_entry.reported_by = request.user
            log_entry.save()
            
            # Handle multiple images
            images = request.FILES.getlist('images')
            for img in images:
                ProjectMonitoringImage.objects.create(
                    monitoring_log=log_entry,
                    image=img
                )

            log_project_activity(
                project=project,
                user=request.user,
                action_type='MONITORING',
                title=f"Site Monitoring Logged ({log_entry.reported_execution_percentage}%)",
                description=f"Engineer assessment: {log_entry.reported_execution_percentage}% complete. Site Notes: {log_entry.description}"
            )
            messages.success(request, f"Monitoring log updated. Project progress set to {log_entry.reported_execution_percentage}%.")
        else:
            messages.error(request, "Error creating monitoring log. Please verify details.")
        return redirect('projects:project_detail', pk=project.pk)


class ProjectMonitoringDashboardView(LoginRequiredMixin, View):
    template_name = 'projects/monitoring_dashboard.html'

    def get(self, request):
        project_id = request.GET.get('project')
        logs_qs = ProjectMonitoringLog.objects.all().prefetch_related('images').select_related('project', 'reported_by')
        selected_project = None
        if project_id and project_id != '0':
            logs_qs = logs_qs.filter(project_id=project_id)
            selected_project = get_object_or_404(Project, pk=project_id)

        context = {
            'logs': logs_qs.order_by('-reported_at'),
            'projects': Project.objects.all(),
            'selected_project': selected_project,
            'monitoring_form': ProjectMonitoringLogGlobalForm(),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        if not (request.user.is_superuser or request.user.groups.filter(name__in=['Level 2', 'Level 3', 'Level 4']).exists()):
            messages.error(request, "You do not have permission to submit progress logs.")
            return redirect('projects:monitoring_dashboard')

        form = ProjectMonitoringLogGlobalForm(request.POST, request.FILES)
        if form.is_valid():
            log_entry = form.save(commit=False)
            log_entry.reported_by = request.user
            log_entry.save()
            
            # Handle multiple images
            images = request.FILES.getlist('images')
            for img in images:
                ProjectMonitoringImage.objects.create(
                    monitoring_log=log_entry,
                    image=img
                )

            log_project_activity(
                project=log_entry.project,
                user=request.user,
                action_type='MONITORING',
                title=f"Site Monitoring Logged ({log_entry.reported_execution_percentage}%)",
                description=f"Engineer assessment: {log_entry.reported_execution_percentage}% complete. Site Notes: {log_entry.description}"
            )
            messages.success(request, f"Monitoring log submitted for project {log_entry.project.project_code}. Progress updated to {log_entry.reported_execution_percentage}%.")
        else:
            messages.error(request, "Failed to submit monitoring log. Please correct form errors.")
        return redirect('projects:monitoring_dashboard')


class SubcontractorPaymentTrancheCreateView(Level3RequiredMixin, View):
    def post(self, request, allocation_pk):
        allocation = get_object_or_404(ProjectAllocation, pk=allocation_pk)
        form = SubcontractorPaymentTrancheForm(request.POST)
        if form.is_valid():
            tranche = form.save(commit=False)
            tranche.allocation = allocation
            tranche.save()

            log_project_activity(
                project=allocation.project,
                user=request.user,
                action_type='TRANCHE',
                title=f"Tranche Paid: ₦{tranche.amount:,.2f}",
                description=f"Disbursed to {allocation.subcontractor.name} on {tranche.date_paid}. Ref: {tranche.payment_reference or 'N/A'}"
            )
            messages.success(request, f"Recorded payment tranche of ₦{tranche.amount:,.2f} for {allocation.subcontractor.name} successfully!")
        else:
            messages.error(request, "Failed to record payment tranche. Please check form values.")
        return redirect('projects:project_detail', pk=allocation.project.pk)


class ProjectExpenseBreakdownView(LoginRequiredMixin, DetailView):
    model = Project
    template_name = 'projects/project_expense_breakdown.html'
    context_object_name = 'project'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.object
        
        fees = project.fees.select_related('fee_type').all()
        lifecycle_stages = project.lifecycle_stages.all()
        unplanned_expenses = project.unplanned_expenses.select_related('reported_by').all()
        allocations = project.projectallocation_set.select_related('subcontractor').prefetch_related('payment_tranches').all()
        
        context.update({
            'fees': fees,
            'lifecycle_stages': lifecycle_stages,
            'unplanned_expenses': unplanned_expenses,
            'allocations': allocations,
            'page_title': f"Expense Breakdown - {project.project_code}",
        })
        return context


def export_project_expense_breakdown(request, pk):
    """Generates an itemized Excel workbook of all incurred expenses for the project."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    project = get_object_or_404(Project, pk=pk)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Breakdown"
    ws.views.sheetView[0].showGridLines = True

    # Styles
    title_font = Font(name="Arial", size=14, bold=True, color="1F2937")
    section_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=10, bold=True, color="374151")
    regular_font = Font(name="Arial", size=10, color="1F2937")
    bold_font = Font(name="Arial", size=10, bold=True, color="1F2937")

    gold_fill = PatternFill(start_color="BFA12C", end_color="BFA12C", fill_type="solid")
    header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    # Title Block
    ws.append([f"PROJECT FINANCIAL EXPENSE BREAKDOWN: {project.project_code}"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Project Name: {project.project_name} | MDA: {project.mda}"])
    ws.append([])

    # KPI Summary Section
    ws.append(["EXECUTIVE FINANCIAL FOOTPRINT SUMMARY"])
    ws.cell(row=4, column=1).font = section_font
    ws.cell(row=4, column=1).fill = gold_fill

    ws.append(["Metric", "Amount (NGN)"])
    ws.cell(row=5, column=1).font = header_font
    ws.cell(row=5, column=2).font = header_font
    
    summary_data = [
        ("Total Contract Amount", float(project.total_actual_contract_amount)),
        ("Total Project Fees (Deductions)", float(project.total_fees_amount)),
        ("Total Fees Paid to Date", float(project.total_fees_paid)),
        ("Total Lifecycle Stage Costs", float(project.total_lifecycle_expenses)),
        ("Total Unplanned / Ad-Hoc Costs", float(project.total_unplanned_expenses)),
        ("Total Subcontractor Agreed Commitments", float(project.total_subcontractor_commitments)),
        ("Total Subcontractor Disbursed Payouts", float(project.total_subcontractor_paid)),
        ("Consolidated Total Project Expenses", float(project.total_project_expenses)),
        ("Net Projected Margin (Profit/Deficit)", float(project.net_project_margin)),
    ]

    for item, val in summary_data:
        ws.append([item, val])
        r = ws.max_row
        ws.cell(row=r, column=2).number_format = '₦#,##0.00'
        ws.cell(row=r, column=1).font = bold_font if "Total" in item or "Consolidated" in item or "Net" in item else regular_font

    ws.append([])

    # Section 1: Project Fees
    ws.append(["1. PROJECT FEES & DEDUCTIONS"])
    ws.cell(row=ws.max_row, column=1).font = section_font
    ws.cell(row=ws.max_row, column=1).fill = gold_fill

    ws.append(["Fee Type", "Status", "Date Paid", "Payment Reference", "Amount (NGN)"])
    r_hdr = ws.max_row
    for c in range(1, 6):
        ws.cell(row=r_hdr, column=c).font = header_font
        ws.cell(row=r_hdr, column=c).fill = header_fill

    for fee in project.fees.select_related('fee_type').all():
        d_paid = fee.date_paid.strftime('%Y-%m-%d') if fee.date_paid else "-"
        ws.append([fee.fee_type.name, fee.get_status_display(), d_paid, fee.payment_reference or "-", float(fee.amount)])
        ws.cell(row=ws.max_row, column=5).number_format = '₦#,##0.00'

    ws.append([])

    # Section 2: Lifecycle Expenses
    ws.append(["2. WORKFLOW LIFECYCLE STAGE EXPENSES"])
    ws.cell(row=ws.max_row, column=1).font = section_font
    ws.cell(row=ws.max_row, column=1).fill = gold_fill

    ws.append(["Sequence", "Stage Name", "Completion Status", "Completed Date", "Notes/Updates", "Incurred Cost (NGN)"])
    r_hdr = ws.max_row
    for c in range(1, 7):
        ws.cell(row=r_hdr, column=c).font = header_font
        ws.cell(row=r_hdr, column=c).fill = header_fill

    for st in project.lifecycle_stages.all():
        c_date = st.completed_date.strftime('%Y-%m-%d') if st.completed_date else "-"
        ws.append([st.sequence_order, st.stage_name, "Completed" if st.is_completed else "Pending", c_date, st.notes_or_updates or "-", float(st.incurred_cost)])
        ws.cell(row=ws.max_row, column=6).number_format = '₦#,##0.00'

    ws.append([])

    # Section 3: Unplanned Expenses
    ws.append(["3. UNPLANNED & AD-HOC EXPENSES"])
    ws.cell(row=ws.max_row, column=1).font = section_font
    ws.cell(row=ws.max_row, column=1).fill = gold_fill

    ws.append(["Date Incurred", "Description", "Reported By", "Amount (NGN)"])
    r_hdr = ws.max_row
    for c in range(1, 5):
        ws.cell(row=r_hdr, column=c).font = header_font
        ws.cell(row=r_hdr, column=c).fill = header_fill

    for exp in project.unplanned_expenses.select_related('reported_by').all():
        r_by = exp.reported_by.get_full_name() if exp.reported_by else "Staff"
        ws.append([exp.date_incurred.strftime('%Y-%m-%d'), exp.description, r_by, float(exp.amount)])
        ws.cell(row=ws.max_row, column=4).number_format = '₦#,##0.00'

    ws.append([])

    # Section 4: Subcontractor Allocations & Payouts
    ws.append(["4. SUBCONTRACTOR ALLOCATIONS & PAYMENT TRANCHES"])
    ws.cell(row=ws.max_row, column=1).font = section_font
    ws.cell(row=ws.max_row, column=1).fill = gold_fill

    ws.append(["Subcontractor", "Company", "Agreed Cost (NGN)", "Advance Paid (NGN)", "Tranches Total (NGN)", "Total Paid (NGN)"])
    r_hdr = ws.max_row
    for c in range(1, 7):
        ws.cell(row=r_hdr, column=c).font = header_font
        ws.cell(row=r_hdr, column=c).fill = header_fill

    for alloc in project.projectallocation_set.select_related('subcontractor').all():
        ws.append([alloc.subcontractor.name, "-", float(alloc.amount_agreed_with_supplier_contractor), float(alloc.advance_received_by_supplier_contractor), float(sum(t.amount for t in alloc.payment_tranches.all())), float(alloc.total_paid)])
        for col_idx in range(3, 7):
            ws.cell(row=ws.max_row, column=col_idx).number_format = '₦#,##0.00'

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="Expense_Breakdown_{project.project_code}.xlsx"'
    wb.save(response)
    return response


def export_projects_excel(request):
    """
    Generates an Excel spreadsheet (.xlsx) of projects matching the active filters
    (search query 'q', year, mda/agency, category, project_type).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    qs = Project.objects.select_related('category').prefetch_related('fees', 'lifecycle_stages', 'unplanned_expenses', 'projectallocation_set').all()

    # Apply same filtering logic as ProjectListView
    q = request.GET.get('q', '').strip()
    year = request.GET.get('year', '').strip()
    mda = request.GET.get('mda', '').strip()
    category = request.GET.get('category', '').strip()
    project_type = request.GET.get('project_type', '').strip()
    awarded = request.GET.get('awarded', '').strip().lower()

    if q:
        qs = qs.filter(Q(project_code__icontains=q) | Q(project_name__icontains=q))
    if year:
        qs = qs.filter(created_at__year=year)
    if mda:
        qs = qs.filter(mda__icontains=mda)
    if category:
        if category.isdigit():
            qs = qs.filter(category_id=category)
        else:
            if category.upper() == 'CONSTRUCTION':
                qs = qs.filter(Q(category__name__iexact='CONSTRUCTION') | Q(category__name__icontains='Civil'))
            else:
                qs = qs.filter(category__name__iexact=category)
    if project_type:
        qs = qs.filter(project_type=project_type)
    if awarded in ['awarded', 'yes', '1', 'true']:
        qs = qs.filter(
            Q(actual_contract_amount__gt=0) |
            Q(current_phase__in=['POST_AWARD', 'EXECUTION']) |
            (Q(award_letter_and_boq__isnull=False) & ~Q(award_letter_and_boq=''))
        )
    elif awarded in ['pre_award', 'no', '0', 'false']:
        qs = qs.filter(
            Q(actual_contract_amount=0) &
            Q(current_phase='PRE_AWARD') &
            (Q(award_letter_and_boq__isnull=True) | Q(award_letter_and_boq=''))
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects Directory"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    title_font = Font(name="Arial", size=14, bold=True, color="1F2937")
    sub_font = Font(name="Arial", size=10, italic=True, color="4B5563")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    regular_font = Font(name="Arial", size=10, color="1F2937")
    bold_font = Font(name="Arial", size=10, bold=True, color="1F2937")

    header_fill = PatternFill(start_color="BFA12C", end_color="BFA12C", fill_type="solid")
    summary_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    # Title Block
    ws.append(["PROJECTS DIRECTORY FINANCIAL & OPERATIONAL EXPORT"])
    ws.cell(row=1, column=1).font = title_font
    
    filter_info = []
    if year: filter_info.append(f"Year: {year}")
    if mda: filter_info.append(f"Agency/MDA: {mda}")
    if category: filter_info.append(f"Category ID: {category}")
    if project_type: filter_info.append(f"Type: {project_type}")
    if awarded in ['awarded', 'yes', '1', 'true']: filter_info.append("Status: Awarded Projects")
    elif awarded in ['pre_award', 'no', '0', 'false']: filter_info.append("Status: Pre-Award Projects")
    if q: filter_info.append(f"Search: '{q}'")
    
    filter_str = " | ".join(filter_info) if filter_info else "All Projects (Unfiltered)"
    ws.append([f"Filters Applied: {filter_str} | Generated: {now().strftime('%Y-%m-%d %H:%M')}"])
    ws.cell(row=2, column=1).font = sub_font
    ws.append([])

    # Table Headers
    headers = [
        "SN", "Project Code", "Project Name", "MDA / Agency", "Lot", "Location", 
        "Category", "Project Type", "Execution Mode", "Completion %",
        "Budget Amount (NGN)", "Contract Amount (NGN)", "Total Fees (NGN)", 
        "Total Stage Costs (NGN)", "Unplanned Costs (NGN)", "Subcontractor Agreed (NGN)",
        "Consolidated Expenses (NGN)", "Net Projected Margin (NGN)", "Technical Status", "Payment Status"
    ]
    ws.append(headers)
    hdr_row = ws.max_row
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=hdr_row, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx in [1, 9, 10, 19, 20] else "left", vertical="center")

    total_budget_sum = 0.0
    total_contract_sum = 0.0
    total_fees_sum = 0.0
    total_expenses_sum = 0.0
    total_margin_sum = 0.0

    for idx, p in enumerate(qs, start=1):
        cat_name = p.category.name if p.category else "-"
        p_type = p.get_project_type_display() if hasattr(p, 'get_project_type_display') else str(p.project_type or "-")
        e_mode = p.get_execution_mode_display() if hasattr(p, 'get_execution_mode_display') else str(p.execution_mode or "-")
        
        b_amt = float(p.budget_amount or 0.0)
        c_amt = float(p.actual_contract_amount or 0.0)
        f_amt = float(p.total_fees_amount or 0.0)
        s_amt = float(p.total_lifecycle_expenses or 0.0)
        u_amt = float(p.total_unplanned_expenses or 0.0)
        sub_amt = float(p.total_subcontractor_commitments or 0.0)
        tot_exp = float(p.total_project_expenses or 0.0)
        net_marg = float(p.net_project_margin or 0.0)

        total_budget_sum += b_amt
        total_contract_sum += c_amt
        total_fees_sum += f_amt
        total_expenses_sum += tot_exp
        total_margin_sum += net_marg

        row_data = [
            idx, p.project_code, p.project_name, p.mda, p.lot or "-", p.location or "-",
            cat_name, p_type, e_mode, f"{p.execution_level_percentage}%",
            b_amt, c_amt, f_amt, s_amt, u_amt, sub_amt, tot_exp, net_marg,
            p.technical_status or "-", p.payment_status or "-"
        ]
        ws.append(row_data)
        curr_row = ws.max_row
        
        # Formatting numbers
        for num_col in range(11, 19):
            ws.cell(row=curr_row, column=num_col).number_format = '₦#,##0.00'

    # Summary Row at Bottom
    ws.append([])
    summary_row = [
        "TOTALS", "", "", f"Total Records: {qs.count()}", "", "", "", "", "", "",
        total_budget_sum, total_contract_sum, total_fees_sum, "", "", "", total_expenses_sum, total_margin_sum, "", ""
    ]
    ws.append(summary_row)
    sum_r = ws.max_row
    for c_idx in range(1, len(summary_row) + 1):
        cell = ws.cell(row=sum_r, column=c_idx)
        cell.font = bold_font
        cell.fill = summary_fill
        if c_idx in [11, 12, 13, 17, 18]:
            cell.number_format = '₦#,##0.00'

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Projects_Export_{year if year else 'All'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class ProjectActivityLogListView(LoginRequiredMixin, ListView):
    model = ProjectActivityLog
    template_name = 'projects/project_activity_log.html'
    context_object_name = 'activities'
    paginate_by = 30

    def get_queryset(self):
        qs = super().get_queryset().select_related('project', 'user')
        action = self.request.GET.get('action', '').strip()
        project_id = self.request.GET.get('project', '').strip()
        q = self.request.GET.get('q', '').strip()

        if action:
            qs = qs.filter(action_type=action)
        if project_id:
            qs = qs.filter(project_id=project_id)
        if q:
            qs = qs.filter(
                Q(title__icontains=q) | Q(description__icontains=q) | Q(project__project_code__icontains=q) | Q(project__project_name__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action_choices'] = ProjectActivityLog.ACTION_CHOICES
        context['projects'] = Project.objects.all().order_by('project_code')
        context['selected_action'] = self.request.GET.get('action', '')
        context['selected_project'] = self.request.GET.get('project', '')
        context['q'] = self.request.GET.get('q', '')

        # Preserve query string for pagination links
        get_copy = self.request.GET.copy()
        if 'page' in get_copy:
            del get_copy['page']
        context['querystring'] = get_copy.urlencode()

        if context.get('is_paginated'):
            page_obj = context['page_obj']
            context['page_range'] = list(page_obj.paginator.get_elided_page_range(
                number=page_obj.number, on_each_side=1, on_ends=1
            ))
        else:
            context['page_range'] = [1]

        return context


